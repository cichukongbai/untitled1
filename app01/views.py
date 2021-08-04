import pymysql
from django.contrib import auth
from django.contrib.auth.decorators import login_required

from django.contrib.auth import get_user_model

from django.contrib.auth.models import *
from django.db.models import Max
from django.shortcuts import render

# Create your views here.
#视图函数,返回index.html页面
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.shortcuts import render, HttpResponse, redirect
from django.http import HttpResponse
from django.views.decorators.clickjacking import xframe_options_exempt
from django.views.decorators.clickjacking import xframe_options_deny
from django.views.decorators.clickjacking import xframe_options_sameorigin
from openpyxl import Workbook

from app01.myform import RegForm
from django.contrib.auth.decorators import login_required, permission_required
# 注册
# Create your views here.
from app01 import models
from app01.models import UserInfo as User
# 注册
def register(request):
    response_dic = {'status': 100, 'msg': None}
    reg_form = RegForm()
    if request.method == 'POST':
        reg_form = RegForm(request.POST)
        if reg_form.is_valid():
            clean_data = reg_form.cleaned_data
            models.UserInfo.objects.create_user(**clean_data)
            response_dic['msg'] = '注册成功'
            response_dic['url'] = '/login/'
        else:
            # 注册失败,返回错误信息,和状态码
            response_dic['status'] = 101
            response_dic['msg'] = reg_form.errors
        return JsonResponse(response_dic)
    return render(request, 'register.html', {'reg_form': reg_form})

# 登录
def login(request):
    if request.method == 'GET':
        print('111')
        return render(request, 'login.html')
    else:
        print('====================')
        response_dic = {'status': 100, 'msg': None}
        name = request.POST.get('name')
        pwd = request.POST.get('pwd')
        request.session['name'] = name
        print(name)
        user = auth.authenticate(username=name, password=pwd)
        # 校验用户
        if user:
            # 如果有值，登陆成功，要login一下
            auth.login(request, user)
            response_dic['msg'] = '登录成功'
        else:
            response_dic['msg'] = '用户名或密码错误'
            response_dic['status'] = 101
        return JsonResponse(response_dic)


# 登出
@login_required(login_url='/login.html')
def logout(request):
    auth.logout(request)
    # del request.session['handle']
    return redirect('/login/')

# 修改密码
@login_required(login_url='/login.html')
def AlterPassword(request):
    print("======AlterPassword========")
    if request.method == 'GET':
        return render(request, 'AlterPassword.html')
    else:
        user=request.user
        old_pw = request.POST.get('old_pw')
        new_pw = request.POST.get('new_pw')
        ret=user.check_password(old_pw)
        if ret:
            user.set_password(new_pw)
            user.save()
            response_dic = {'status': 100, 'msg': "修改成功！"}
        else:
            response_dic = {'status': 101, 'msg': "修改失败！"}
    return JsonResponse(response_dic)


#欢迎页面
def index(request):
     return render(request, 'index-32.html',{'gglist1': '1'})

def test(request):
     return render(request, 'test.html')




# 权限管理
#仅仅用于用户信息查询
@login_required(login_url='/login.html')
# @permission_required('app01.change_app01_temp_tem')
def PerAdmin(request):
    if request.method == 'GET':
        gg_list = models.UserInfo.objects.all()
        tt_list = Permission.objects.all()
        print(tt_list)
        return render(request, 'PerAdmin.html', {'gglist': gg_list,'ttlist': tt_list})
    else:
        userid = request.POST['id_no']
        print(userid)
        permissions = request.POST['permissions'].split(', ')
        if permissions[0] != '':  # 不为空时
            print(permissions)
            for index,value in enumerate(permissions):
                permissions[index]=value  # 将 数字 替换为 上面数组中的 字符串
            for per in permissions:
                User.objects.get(nid=userid).user_permissions.add(per)
            response_dic = {'status': 100, 'msg': '修改权限成功'}
        else:
            response_dic = {'status': 101, 'msg': '未选择权限'}
        return JsonResponse(response_dic)


# 参数修改
@login_required(login_url='/login.html')
def AlterParameter(request):
    print("AlterParameter")
    if request.method == 'GET':
        return render(request, 'AlterParameter.html')
    else:
        school = request.POST.get('id_name1')
        # 校区
        colle = request.POST.get('id_colle')
        # id_colle=colle
        year = request.POST.get('id_year')
        seame = request.POST.get('id_seame')
        print(school , colle , year , seame)
        # 如何把下拉框中默认的选项去掉
        if school and colle and year and seame:
            if year.isdigit():
                if is_Chinese(school)==True:
                    cid = models.Parameter.objects.filter(id=1).values('id')
                    if cid.exists():
                         models.Parameter.objects.filter(id=1).update(School=school, SYear=year, Semester=seame, Campus_id=colle)
                    else:
                        models.Parameter.objects.create(id=1,School=school, SYear=year, Semester=seame, Campus_id=colle)

                    response_dic = {'status': 100, 'msg': '参数设置成功：'+school+year+seame+colle}
                else:
                    response_dic = {'status': 103, 'msg': '学校应为汉字：'}
            else:
                response_dic = {'status': 102, 'msg': '年份输入有误' }
        else:
            response_dic = {'status': 101, 'msg': '请填写完整！'}
    return JsonResponse(response_dic)






# 教室总表进行数据查询
@login_required(login_url='/login.html')
# @permission_required('app01.can_view_classroom')
def Summary(request):
    if request.method == 'GET':
        ll_list=models.Campus.objects.all()
        cc_list=models.ClassroomType.objects.all()
        gg_list = models.Arrangement.objects.all()
        return render(request, 'Summary.html', {'gglist': gg_list,'lllist': ll_list,'cclist': cc_list})
    else:
        flag=int(request.POST.get('flag'))
        print(flag)
        type = request.POST.get('id_type')

        camp = request.POST.get('id_camp')
        print(type,camp)
        # 查询
        if flag==1:
            print("Summary type  camp", type, camp)
            # request.session['flag'] = flag
            # request.session['camp'] = camp
            # if camp == '全部':
            #     if type == '1':
            #         gglist = models.Arrangement.objects.all().order_by('RoomID_id')
            #     else:
            #         gglist = models.Arrangement.objects.all().order_by('Enroll')
            # else:
            # 按教室编号
            if type == '1':
                gglist = models.Arrangement.objects.filter(campus_id=camp).order_by('RoomID_id')
            # 按容量
            else:
                gglist = models.Arrangement.objects.filter(campus_id=camp).order_by('Enroll')
            return render(request, 'Summary.html', context={'gglist': gglist})
        # 打印
        if flag == 2:
            print('开始打印')
            wb = Workbook()  # optimized_write=True
            sheet_name = '排课总表'
            if camp == '全部':
                if type == '1':
                    basic_obj = models.Arrangement.objects.all().order_by('RoomID_id')
                else:
                    basic_obj = models.Arrangement.objects.all().order_by('Enroll')
            else:
                if type == '1':
                    basic_obj = models.Arrangement.objects.filter(campus_id=camp).order_by('RoomID_id')
                else:
                    basic_obj = models.Arrangement.objects.filter(campus_id=camp).order_by('Enroll')
            w1 = wb.create_sheet('基本信息', 0)
            name_list1 = ['教室名称', '容量', '星期一 单/双','星期二 单/双','星期二 单/双','星期二 单/双','星期二 单/双','星期二 单/双','星期二 单/双','星期二 单/双','星期三 单/双','星期三 单/双','星期三 单/双','星期三 单/双','星期三 单/双','星期三 单/双','星期三 单/双','星期四 单/双','星期四 单/双','星期四 单/双','星期四 单/双','星期四 单/双','星期四 单/双','星期四 单/双','星期五 单/双','星期五 单/双','星期五 单/双','星期五 单/双','星期五 单/双','星期五 单/双','星期五 单/双','星期六 单/双','星期六 单/双','星期六 单/双','星期六 单/双','星期六 单/双','星期六 单/双','星期六 单/双','星期日 单/双']
            name_list2=['教室名称','容量','1-2','3-4','5-6','7-8','9-10','11-13','1-2','3-4','5-6','7-8','9-10','11-13','1-2','3-4','5-6','7-8','9-10','11-13','1-2','3-4','5-6','7-8','9-10','11-13','1-2','3-4','5-6','7-8','9-10','11-13','1-2','3-4','5-6','7-8','9-10','11-13','1-2','3-4','5-6','7-8','9-10','11-13']
            for i in range(1, len(name_list1) + 1):
                w1.cell(row=1, column=i, value=name_list1[i - 1])
            for j in range(1, len(name_list2) + 1):
                w1.cell(row=2, column=j, value=name_list2[j - 1])

            ws = wb.active
            ws.merge_cells('C1:H1')
            ws.merge_cells('I1:N1')
            ws.merge_cells('O1:T1')
            ws.merge_cells('U1:Z1')
            ws.merge_cells('AA1:AF1')
            ws.merge_cells('AG1:AL1')
            ws.merge_cells('AM1:AR1')
            # ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=8)
            # ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=9)
            # ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=10)
            # ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=11)
            # ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=12)
            # ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=13)
            # ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=14)
            ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
            ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
            excel_row = 3
            # 导入mysql中的数据
            for obj in basic_obj:
                roomID = obj.RoomID_id
                enroll = obj.Enroll
                Mon1_2=obj.Mon1_2
                Mon3_4=obj.Mon3_4
                Mon5_6=obj.Mon5_6
                Mon7_8=obj.Mon7_8
                Mon9_10=obj.Mon9_10
                Mon11_13=obj.Mon11_13
                Tue1_2 = obj.Tue1_2
                Tue3_4 = obj.Tue3_4
                Tue5_6 = obj.Tue5_6
                Tue7_8 = obj.Tue7_8
                Tue9_10 = obj.Tue9_10
                Tue11_13 = obj.Tue11_13
                Wed1_2 = obj.Wed1_2
                Wed3_4 = obj.Wed3_4
                Wed5_6 = obj.Wed5_6
                Wed7_8 = obj.Wed7_8
                Wed9_10 = obj.Wed9_10
                Wed11_13 = obj.Wed11_13
                Thu1_2 = obj.Thu1_2
                Thu3_4 = obj.Thu3_4
                Thu5_6 = obj.Thu5_6
                Thu7_8 = obj.Thu7_8
                Thu9_10 = obj.Thu9_10
                Thu11_13 = obj.Thu11_13
                Fri1_2 = obj.Fri1_2
                Fri3_4 = obj.Fri3_4
                Fri5_6 = obj.Fri5_6
                Fri7_8 = obj.Fri7_8
                Fri9_10 = obj.Fri9_10
                Fri11_13 = obj.Fri11_13
                Sat1_2 = obj.Sat1_2
                Sat3_4 = obj.Sat3_4
                Sat5_6 = obj.Sat5_6
                Sat7_8 = obj.Sat7_8
                Sat9_10 = obj.Sat9_10
                Sat11_13 = obj.Sat11_13
                Sun1_2 = obj.Sun1_2
                Sun3_4 = obj.Sun3_4
                Sun5_6 = obj.Sun5_6
                Sun7_8 = obj.Sun7_8
                Sun9_10 = obj.Sun9_10
                Sun11_13 = obj.Sun11_13
                obj_list = [roomID, enroll,Mon1_2,Mon3_4,Mon5_6,Mon7_8,Mon9_10,Mon11_13,Tue1_2,Tue3_4,Tue5_6,Tue7_8,Tue9_10,Tue11_13,Wed1_2,Wed3_4,Wed5_6,Wed7_8,Wed9_10,Wed11_13,Thu1_2,Thu3_4,Thu5_6,Thu7_8,Thu9_10,Thu11_13,Fri1_2,Fri3_4,Fri5_6,Fri7_8,Fri9_10,Fri11_13,Sat1_2,Sat3_4,Sat5_6,Sat7_8,Sat9_10,Sat11_13,Sun1_2,Sun3_4,Sun5_6,Sun7_8,Sun9_10,Sun11_13]
                for i in range(1, len(obj_list)):
                    w1.cell(row=excel_row, column=i, value=obj_list[i - 1])
                excel_row += 1
            print('ok!')
            response = HttpResponse(content_type='application/msexcel')
            response['Content-Disposition'] = 'attachment;filename=%s.xlsx' % sheet_name
            response['status'] = 100
            wb.save('排课.xlsx')
            wb.save(response)
            return response


#获取前端传来的教师总表数据
# @login_required(login_url='/login.html')
#
# def Summary_1(request):
#
#     print("====Summary_1====")
#     if request.method == 'GET':
#         print("GET")
#         return render(request,'Summary.html')
#     if request.method == 'POST':
#         id_Sumtype = request.POST.get('id_type')
#         id_Sumcamp = request.POST.get('id_camp')
#         request.session['id_Sumtype'] = id_Sumtype
#         request.session['id_Sumcamp'] = id_Sumcamp
#         print("Summary_1 type  camp",id_Sumtype,id_Sumcamp)
#         response_dic = {'status': 101, 'msg': '请输入全部内容！'}
#         return JsonResponse(response_dic)
#                     #1代表教室编号
#                     #2代表容量

#用来出处理查询信息
@login_required(login_url='/login.html')
def Summary1(request):

    print('====Summary1====')
    id3 = request.session.get('flag')
    camp=request.session.get('camp')

    print(id3)
    if camp == '全部':
        if id3 == '1':
            gglist = models.Arrangement.objects.all()
            return render(request, 'Summary.html', context={'gglist': gglist})
        elif id3 == '2':
            print('111')
            gglist = models.Arrangement.objects.all().order_by('Enroll')
            print(gglist)
            return render(request, 'Summary.html', context={'gglist': gglist})
    else:
        if id3 == '1':
            gglist = models.Arrangement.objects.filter(campus_id=camp)
            return render(request, 'Summary.html', context={'gglist': gglist})
        elif id3 == '2':
            print('111')
            gglist = models.Arrangement.objects.filter(campus_id=camp).order_by('Enroll')
            print(gglist)
            return render(request, 'Summary.html', context={'gglist': gglist})




# 教室简表
@login_required(login_url='/login.html')
def Simple1(request):
    return render(request, 'Simple.html')
@login_required(login_url='/login.html')
def Simple_1(request):
    print('======Simple_1函数用来存放数据========')

    if request.method == 'GET':
        return render(request, 'Simple.html')
    if request.method == 'POST':
        sim_flag = request.POST.get('sim_flag')
        sim_camp = request.POST.get('sim_camp')
        print('Simple_1  sim_flag sim_camp',sim_flag,sim_camp)

        request.session['sim_flag'] = sim_flag
        request.session['sim_camp'] = sim_camp
        response_dic = {'status': 101, 'msg': None}
        if sim_flag == '2':
            response_dic = {'status': 101, 'msg': None}
            id21 = request.POST.get('id21')
            id22 = request.POST.get('id22')
            print('教室编号范围：', id21, id22)
            if id21 and id22:
                request.session['id21'] = id21
                request.session['id22'] = id22
            else:
                request.session['sim_flag'] = '1'
            return JsonResponse(response_dic)
        return JsonResponse(response_dic)

@login_required(login_url='/login.html')
def Simple(request):
    gg_list = models.Arrangement.objects.all()
    return render(request, 'Simple.html', {'gglist': gg_list})
    # print("=====Simple函数======")
    # sim_flag = request.session.get('sim_flag')
    # sim_camp=request.session.get('sim_camp')
    # id1 = request.session.get('id21')
    # id2 = request.session.get('id22')
    # name = request.session.get('name', default='')
    # print('获得参数sim_flag, sim_camp,id1,id2,name', sim_flag, sim_camp,id1,id2,name)
    # if sim_camp==4:
    #      if sim_flag=='1':
    #         gg_list = models.Arrangement.objects.all()
    #         print('all')
    #                 # print(gg_list)
    #         return render(request, 'Simple.html', {'gglist': gg_list})
    #      elif sim_flag=='2':
    #                 # print(id1)
    #                 # print(type(id1))
    #         gg_list = models.Arrangement.objects.filter(RoomID_id__gte=int(id1), RoomID_id__lte=int(id2)).all()
    #         return render(request, 'Simple.html', {'gglist': gg_list})
    # else:
    #     if sim_flag=='1':
    #         gg_list = models.Arrangement.objects.filter(campus_id=sim_camp)
    #                 # print(gg_list)
    #         return render(request, 'Simple.html', {'gglist': gg_list})
    #     elif sim_flag=='2':
    #                 # print(id1)
    #                 # print(type(id1))
    #         gg_list = models.Arrangement.objects.filter(campus_id=sim_camp,RoomID_id__gte=int(id1), RoomID_id__lte=int(id2)).all()
    #         return render(request, 'Simple.html', {'gglist': gg_list})

def is_Chinese(word):
    for ch in word:
        if '\u4e00' <= ch <= '\u9fff':
            return True
    return False

# # 课程处理
@login_required(login_url='/login.html')
def HandleClass(request):

    if request.method == 'GET':
        return render(request, 'HandleClass.html')

    else:
        request.session['handle'] = True
        camp=request.POST.get('id_camp')

        if camp == '宝山':
            response_dic = {'status': 100, 'msg': '宝山处理完毕'}
            ab = models.Cou.objects.filter(TIMETEXT__contains='上机',campus='宝山')
            for a in ab:
                print(a.coID)
                f = list(a.TIMESET)
                print(a.TIMETEXT.split(" "))
                b = a.TIMETEXT.split(" ")[-2]
                c = int(b.split("-")[0][-1])
                week = b.split("-")[0][-2]
                d = int(b.split("-")[1][0])
                # print(b.split("-")[0][-1])
                print(type(c))
                sum = 13
                if week == '一':
                    sum = sum * 0
                elif week == '二':
                    sum = sum * 1
                elif week == '三':
                    sum = sum * 2
                elif week == '四':
                    sum = sum * 3
                elif week == '五':
                    sum = sum * 4
                elif week == '六':
                    sum = sum * 5
                elif week == '日':
                    sum = sum * 6
                for i in range(c, d + 1):
                    set = sum + i - 1
                    f[set] = '0'
                re = ''.join(f)
                models.Cou.objects.filter(coID=a.coID,campus='宝山').update(TIMESET1=re)
            return JsonResponse(response_dic)
        elif camp == '嘉定':
            response_dic = {'status': 101, 'msg': '嘉定处理完毕'}
            ab = models.Cou.objects.filter(TIMETEXT__contains='上机',campus='嘉定')
            for a in ab:
                print(a.coID)
                f = list(a.TIMESET)
                print(a.TIMETEXT.split(" "))
                b = a.TIMETEXT.split(" ")[-2]
                c = int(b.split("-")[0][-1])
                week = b.split("-")[0][-2]
                d = int(b.split("-")[1][0])
                # print(b.split("-")[0][-1])
                print(type(c))
                sum = 13
                if week == '一':
                    sum = sum * 0
                elif week == '二':
                    sum = sum * 1
                elif week == '三':
                    sum = sum * 2
                elif week == '四':
                    sum = sum * 3
                elif week == '五':
                    sum = sum * 4
                elif week == '六':
                    sum = sum * 5
                elif week == '日':
                    sum = sum * 6
                for i in range(c, d + 1):
                    set = sum + i - 1
                    f[set] = '0'
                re = ''.join(f)
                models.Cou.objects.filter(coID=a.coID,campus='嘉定').update(TIMESET1=re)
            return JsonResponse(response_dic)
        elif camp == '延长':
            response_dic = {'status': 102, 'msg': '延长处理完毕'}
            ab = models.Cou.objects.filter(TIMETEXT__contains='上机',campus='延长')
            for a in ab:
                print(a.coID)
                f = list(a.TIMESET)
                print(a.TIMETEXT.split(" "))
                b = a.TIMETEXT.split(" ")[-2]
                c = int(b.split("-")[0][-1])
                week = b.split("-")[0][-2]
                d = int(b.split("-")[1][0])
                # print(b.split("-")[0][-1])
                print(type(c))
                sum = 13
                if week == '一':
                    sum = sum * 0
                elif week == '二':
                    sum = sum * 1
                elif week == '三':
                    sum = sum * 2
                elif week == '四':
                    sum = sum * 3
                elif week == '五':
                    sum = sum * 4
                elif week == '六':
                    sum = sum * 5
                elif week == '日':
                    sum = sum * 6
                for i in range(c, d + 1):
                    set = sum + i - 1
                    f[set] = '0'
                re = ''.join(f)
                models.Cou.objects.filter(coID=a.coID,campus='延长').update(TIMESET1=re)
            return JsonResponse(response_dic)
        elif camp == '全部':
            response_dic = {'status': 103, 'msg': '全部处理完毕'}
            ab = models.Cou.objects.filter(TIMETEXT__contains='上机')
            for a in ab:
                print(a.coID)
                f = list(a.TIMESET)
                print(a.TIMETEXT.split(" "))
                b = a.TIMETEXT.split(" ")[-2]
                c = int(b.split("-")[0][-1])
                week = b.split("-")[0][-2]
                d = int(b.split("-")[1][0])
                # print(b.split("-")[0][-1])
                print(type(c))
                sum = 13
                if week == '一':
                    sum = sum * 0
                elif week == '二':
                    sum = sum * 1
                elif week == '三':
                    sum = sum * 2
                elif week == '四':
                    sum = sum * 3
                elif week == '五':
                    sum = sum * 4
                elif week == '六':
                    sum = sum * 5
                elif week == '日':
                    sum = sum * 6
                for i in range(c, d + 1):
                    set = sum + i - 1
                    f[set] = '0'
                re = ''.join(f)
                models.Cou.objects.filter(coID=a.coID).update(TIMESET1=re)
            return JsonResponse(response_dic)



def EditStudent(request):
    return render(request,'edit-student.html')


def allprofessors(request):
    return render(request, 'ClassView1.html')